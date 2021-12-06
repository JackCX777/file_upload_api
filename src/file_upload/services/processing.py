import os
from time import sleep

# from django.conf import settings

from xlrd import open_workbook
from openpyxl import load_workbook


TARGET_HEADERS = {
    'before',
    'after'
}


def get_file_extension(file_path: str) -> str:
    """
    Required to determine which library to work with.
    """
    filename, extension = os.path.splitext(file_path)
    return extension.lstrip('.')


def get_target_columns_xls(path: str) -> dict:
    """
    Retrieves the required columns of values from the xls file
    """
    target_columns = {}
    workbook = open_workbook(path)

    for sheet in workbook.sheet_names():
        current_sheet = workbook[sheet]
        header_rows = [current_sheet.row_values(i) for i in range(2)]

        for header_row_indx, header_row in enumerate(header_rows):

            for header_col_indx, header in enumerate(header_row):
                header = header.lower()

                if header in TARGET_HEADERS:
                    data = current_sheet.col_values(header_col_indx, header_row_indx + 1)
                    target_columns[header] = [int(val) for val in data if val]

    return target_columns


def get_target_columns_xlsx(path: str) -> dict:
    """
    Retrieves the required columns of values from the xlsx file
    """
    target_columns = {}
    workbook = load_workbook(path, read_only=True)

    for sheet in workbook.sheetnames:
        current_sheet = workbook[sheet]
        header_rows = [row for row in current_sheet.iter_rows(min_row=1, max_row=2, values_only=True)]

        for header_row_indx, header_row in enumerate(header_rows):
            header_row_indx += 1

            for header_col_indx, header in enumerate(header_row):
                if header is not None:
                    header = str(header.lower())

                if header in TARGET_HEADERS:
                    data = []
                    for row in current_sheet.iter_rows(
                            min_row=header_row_indx + 1,
                            max_row=current_sheet.max_row,
                            min_col=header_col_indx + 1,
                            max_col=header_col_indx + 1,
                            values_only=True
                    ):
                        for val in row:
                            if val is not None:
                                data.append(int(val))

                    target_columns[header] = data

    return target_columns


# Extensions are defined here
GET_TARGET_COLUMNS = {
    'xls': get_target_columns_xls,
    'xlsx': get_target_columns_xlsx
}


def get_target_columns(path: str) -> callable:
    """
    Calls a function suitable for expansion
    """
    return GET_TARGET_COLUMNS.get(get_file_extension(path))(path)


def get_result(path: str) -> str:
    """
    Gets the required final result.
    """
    columns = get_target_columns(path)
    sleep(30)
    if len(columns.get('before')) < len(columns.get('after')):
        for val in columns.get('after'):
            if val not in set(columns.get('before')):
                return f'added: {val}'
    else:
        for val in columns.get('before'):
            if val not in set(columns.get('after')):
                return f'removed: {val}'
